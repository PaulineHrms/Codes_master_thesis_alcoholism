# -*- coding: utf-8 -*-
"""
Created on Tue Mar 21 09:53:24 2023

@author: Pauline
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import scipy.stats as stats
import os
from openpyxl import Workbook
from path import perso_path_string

plt.rcParams["figure.figsize"] = (5,5)
perso_path, excel_path, subjects_path, patients_path, analysis_path, atlas_path, P_folder_path, Freesurfer_path = perso_path_string(on_cluster=False)
Test_evol_PATH = P_folder_path + "Analysis/Test_evolution_DIFF/"


## RENAME VERSION ## + OUTLIERS
region_list = ["CC", "CC_genu", "CC_ant_midbody", "CC_post_midbody", "CC_isthmus", "CC_splenium", "UF_left", "UF_right", "UF"]
models = ["dti","diamond", "mf", "noddi"]
metric_list = {"dti" : ["FA", "MD", "AD", "RD"],"diamond" : ["wFA", "wMD", "wAD", "wRD" , "frac_csf" , "frac_ftot"],"mf" : ["fvf_tot", "frac_csf", "frac_ftot", "wfvf"],"noddi" : ["fiso", "fintra", "fextra", "odi"]}
comportment_list = ["BDI", "Total_OCDS", "OCDS_Obsessions", "OCDS_Compulsions", "STAI","MFI"]
time = ["T1", "T2"]
Stream_list = ["nb_stream", "mean_streamDensity", "std_streamDensity"]

datas_c = pd.read_excel(excel_path + "data_alcoholic_controls_rename_without_outliers.xlsx",index_col=0 ).replace(["/", "perdu", "incomplet"], np.nan)
datas_p = pd.read_excel(excel_path + "data_alcoholic_patients_rename_without_outliers.xlsx",index_col=0 ).replace(["/", "perdu", "incomplet"], np.nan)

DATA = {"Patients" : datas_p, "Controls" : datas_c}



for cible in ["Patients", "Controls"]:
    for model in models:
        if not os.path.exists(Test_evol_PATH + "/"+ cible + "/" + model + "/" ): 
            os.mkdir(Test_evol_PATH + "/" + cible + "/" + model + "/")
        for metric in metric_list[model]:
            if not os.path.exists(Test_evol_PATH + "/" + cible + "/" + model + "/" + metric): 
                os.mkdir(Test_evol_PATH + "/" + cible + "/" + model + "/" + metric)

if not os.path.exists(Test_evol_PATH + "/Patients/Comportement/"): 
    os.mkdir(Test_evol_PATH + "/Patients/Comportement/")
    
if not os.path.exists(Test_evol_PATH + "/Patients/Test_evolution_pval_patients.xlsx")  :
    wb = Workbook()
    wb.save(Test_evol_PATH + "/Patients/Test_evolution_pval_patients.xlsx")
    
if not os.path.exists(Test_evol_PATH + "/Controls/Test_evolution_pval_controls.xlsx")  :
    wb = Workbook()
    wb.save(Test_evol_PATH + "/Controls/Test_evolution_pval_controls.xlsx")
    
if not os.path.exists(Test_evol_PATH + "/Test_evolution.xlsx")  :
    wb = Workbook()
    wb.save(Test_evol_PATH + "/Test_evolution.xlsx")
#%%

def significative_diff(p, c):
    if p < 0 :
        if c < 0 : 
            if p < c : return True
            else : return False
        if c > 0 : return True
    else :  # p > 0
        if c > 0 : 
            if p > c : return True
            else : return False
        else : return True
    
            
#%%
#cible = "Patients"
#models = ["dti"]
#region_list = ["CC", "UF"]


df_pval_c = pd.DataFrame(region_list, columns=(["Region"]))
df_mean_c = pd.DataFrame(region_list, columns=(["Region"]))

df_pval_p = pd.DataFrame(region_list, columns=(["Region"]))
df_mean_p = pd.DataFrame(region_list, columns=(["Region"]))

df_diff = pd.DataFrame(region_list, columns=(["Region"]))

for model in models:
    for metric in metric_list[model]:
        for region in region_list:
            
            data_p = (datas_p[["wMean_"+ metric + "_"+ model+"_T1_"+region, "wMean_"+ metric + "_"+ model+"_T2_"+region]]).copy().dropna(axis = 0)
            data_c = (datas_c[["wMean_"+ metric + "_"+ model+"_T1_"+region, "wMean_"+ metric + "_"+ model+"_T2_"+region]]).copy().dropna(axis = 0)
            
            #for controls
            data_T1_c = np.asarray(data_c["wMean_"+ metric + "_"+ model+"_T1_"+region])
            data_T2_c = np.asarray(data_c["wMean_"+ metric + "_"+ model+"_T2_"+region])
            
            diff_c = data_T2_c - data_T1_c
            Mean_c = np.mean(diff_c)
            df_mean_c.loc[df_mean_c["Region"] == region,[model + "_" + metric]] = Mean_c
            
            t_Score_c , p_val_c = stats.ttest_rel(data_T1_c, data_T2_c)
            df_pval_c.loc[df_pval_c["Region"] == region,[model + "_" + metric]] = round(p_val_c, 4)
            
            a = """
            fig, ax = plt.subplots()            
            labels = ['T1', 'T2']
            sns.violinplot([data_T1_c, data_T2_c],showmeans=True,showmedians=False, palette = ["darkseagreen","cadetblue"] )
            ax.yaxis.grid(True)
            ax.set_xticks([y for y in range(2)], labels=['T1', 'T2'])
            plt.title("Evolution " + model + " " + metric + " in " + region + " (Controls) \n Ttest p-val= "+ str(round(p_val_c, 4)))
            x = [0,1]
            for i in range(len(data_T1_c)):
                y = [data_T1_c[i], data_T2_c[i]]
                plt.plot(x,y, color="sandybrown", alpha= 0.5)
            plt.savefig(Test_evol_PATH+ "/"+ cible + "/" + model + "/" + metric + "/Evolution " + model + "-" + metric + "-" + region + ".png")
            plt.show()"""
            
            #for patients
            data_T1_p = np.asarray(data_p["wMean_"+ metric + "_"+ model+"_T1_"+region])
            data_T2_p = np.asarray(data_p["wMean_"+ metric + "_"+ model+"_T2_"+region])
            
            diff_p = data_T2_p - data_T1_p
            Mean_p = np.mean(diff_p)
            df_mean_p.loc[df_mean_p["Region"] == region,[model + "_" + metric]] = Mean_p
            
            if significative_diff(Mean_p, Mean_c) : 
                df_diff.loc[df_diff["Region"] == region,[model + "_" + metric]] = "OK"
                print(metric, model, region, "diff significative")
            else :
                df_diff.loc[df_diff["Region"] == region,[model + "_" + metric]] = "NON sign"
                print(metric, model, region, "diff NON significative")
            
            t_Score_p , p_val_p = stats.ttest_rel(data_T1_p, data_T2_p)
            df_pval_p.loc[df_pval_p["Region"] == region,[model + "_" + metric]] = round(p_val_p, 4)
            
            b = """
            fig, ax = plt.subplots()            
            labels = ['T1', 'T2']
            sns.violinplot([data_T1_p, data_T2_p],showmeans=True,showmedians=False, palette = ["darkseagreen","cadetblue"] )
            ax.yaxis.grid(True)
            ax.set_xticks([y for y in range(2)], labels=['T1', 'T2'])
            plt.title("Evolution " + model + " " + metric + " in " + region + " (Patients) \n Ttest p-val= "+ str(round(p_val_p, 4)))
            x = [0,1]
            for i in range(len(data_T1_p)):
                y = [data_T1_p[i], data_T2_p[i]]
                plt.plot(x,y, color="sandybrown", alpha= 0.5)
            plt.savefig(Test_evol_PATH+ "/"+ cible + "/" + model + "/" + metric + "/Evolution " + model + "-" + metric + "-" + region + ".png")
            plt.show()"""

with pd.ExcelWriter(Test_evol_PATH+ "/Patients/Test_evolution_pval_patients.xlsx",
                engine='openpyxl',mode='a', if_sheet_exists = 'replace') as writer:  
    df_pval_p.to_excel(writer, sheet_name="Metrics_pval")
    df_mean_p.to_excel(writer, sheet_name="Metrics_mean")
    
with pd.ExcelWriter(Test_evol_PATH+ "/Controls/Test_evolution_pval_controls.xlsx",
                engine='openpyxl',mode='a', if_sheet_exists = 'replace') as writer:  
    df_pval_c.to_excel(writer, sheet_name="Metrics_pval")
    df_mean_c.to_excel(writer, sheet_name="Metrics_mean")

with pd.ExcelWriter(Test_evol_PATH+ "/Test_evolution.xlsx",
                engine='openpyxl',mode='a', if_sheet_exists = 'replace') as writer:  
    df_pval_c.to_excel(writer, sheet_name="Metrics_pval_c")
    df_mean_c.to_excel(writer, sheet_name="Metrics_mean_c")
    df_pval_p.to_excel(writer, sheet_name="Metrics_pval_p")
    df_mean_p.to_excel(writer, sheet_name="Metrics_mean_p")
    df_diff.to_excel(writer, sheet_name="Sign._diff")
    

#%%

df_pval = pd.DataFrame(comportment_list, columns=(["Behavioural metrics"]))

cible = "Patients" #Pas de T2 dans controls donc il faut jamais changer

datas = DATA[cible]
for c in comportment_list :
        
        data = (datas[["Numéro","T1_"+c, "T2_"+c]]).copy().dropna(axis = 0)
        data_T1 = np.asarray(data["T1_"+c])
        data_T2 = np.asarray(data["T2_"+c])
        ttest = stats.ttest_rel(data_T1, data_T2)
        t_score, p_val = ttest

        df_pval.loc[df_pval["Behavioural metrics"] == c,["p_val"]] = format(p_val, ".1e")
         
        
        diff = data_T2 - data_T1
        Mean = np.mean(diff)
        df_pval.loc[df_pval["Behavioural metrics"] == c,["mean"]] = Mean
        
        
        fig, ax = plt.subplots()
        labels = ['T1', 'T2']
        sns.violinplot([data_T1, data_T2],showmeans=True,
                      showmedians=False,palette = ["darkseagreen","cadetblue"])

                
        ax.yaxis.grid(True)
        ax.set_xticks([y for y in range(2)], labels=['T1', 'T2'])

        plt.title("Evolution " +c+ "\n Ttest p-val = "+ str(format(p_val, ".1e")))
        data_T1 = np.asarray(data["T1_"+c])
        data_T2 = np.asarray(data["T2_"+c])

        x = [0,1]
        for i in range(len(data_T1)):
            y = [data_T1[i], data_T2[i]]
            ax.plot(x,y,color = "sandybrown", alpha = 0.5)
        plt.savefig(Test_evol_PATH+ "/"+ cible + "/"+ "Comportement/Evolution " +c + ".png")
        plt.show()


with pd.ExcelWriter(Test_evol_PATH+ "/"+ cible + "/Test_evolution_pval_" + cible + ".xlsx",
                engine='openpyxl',mode='a', if_sheet_exists = 'replace') as writer:  
    df_pval.to_excel(writer, sheet_name="Comportment_pval")
    
    
#%%

if False : 
    cible = "Patients"
    datas = DATA[cible]
    for model in models:
        for metric in metric_list[model]:
            for region in region_list:
                #print(r)
                
                data = (datas[["wMean_"+ metric + "_"+ model+"_T1_"+region,"wMean_"+ metric + "_"+ model+"_T2_"+region]]).copy().dropna(axis = 0)
                #print(dic["data_"+r].shape)
                
                data_T1 = np.asarray(data["wMean_"+ metric + "_"+ model+"_T1_"+region])
                data_T2 = np.asarray(data["wMean_"+ metric + "_"+ model+"_T2_"+region])
                
                x = ["T1", "T2"]
                
                for i in range(len(data_T1)):
                    y = [data_T1[i], data_T2[i]]
                    plt.plot(x,y)
                plt.title("Evolution " + model + " " + metric + " in the " + region)
                plt.show()
    
    for c in comportment_list : 
        data = (datas[["Numéro","T1_"+c, "T2_"+c]]).copy().dropna(axis = 0)
    
        data_E1 = np.asarray(data["T1_"+c])
        data_E2 = np.asarray(data["T2_"+c])
        
        x = ["T1", "T2"]
        for i in range(len(data_T1)):
            y = [data_T1[i], data_T2[i]]
            plt.plot(x,y, label = i)
        plt.title("Evolution " +c)
        #plt.legend()
        plt.show()  
        