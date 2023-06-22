# -*- coding: utf-8 -*-
"""
Created on Tue May 30 15:28:49 2023

@author: Pauline
"""


import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import scipy.stats as stats
import os
from openpyxl import Workbook
from path import perso_path_string

plt.rcParams["figure.figsize"] = (5,5)
perso_path, excel_path, subjects_path, patients_path, analysis_path, atlas_path, P_folder_path, Freesurfer_path = perso_path_string(on_cluster=False)
Test_evol_PATH = P_folder_path + "Analysis/Test_evolution/DIFF/"


## RENAME VERSION ## + OUTLIERS
region_list = ["CC", "CC_genu", "CC_ant_midbody", "CC_post_midbody", "CC_isthmus", "CC_splenium", "UF_left", "UF_right", "UF"]
models = ["dti","diamond", "mf", "noddi"]
metric_list = {"dti" : ["FA", "MD", "AD", "RD"],
               "diamond" : ["wFA", "wMD", "wAD", "wRD" , "frac_csf" , "frac_ftot"],
               "mf" : ["fvf_tot", "frac_csf", "frac_ftot", "wfvf"],
               "noddi" : ["fiso", "fintra", "fextra", "odi"]}
comportment_list = ["BDI", "Total_OCDS", "OCDS_Obsessions", "OCDS_Compulsions", "STAI","MFI"]
time = ["T1", "T2"]
Stream_list = ["nb_stream", "mean_streamDensity", "std_streamDensity"]

datas_c = pd.read_excel(excel_path + "data_alcoholic_controls_rename_without_outliers.xlsx",index_col=0 ).replace(["/", "perdu", "incomplet"], np.nan)
datas_p = pd.read_excel(excel_path + "data_alcoholic_patients_rename_without_outliers.xlsx",index_col=0 ).replace(["/", "perdu", "incomplet"], np.nan)

DATA = {"Patients" : datas_p, "Controls" : datas_c}


a  = """
for cible in ["Patients", "Controls"]:
    for model in models:
        if not os.path.exists(Test_evol_PATH + "/"+ cible + "/" + model + "/" ): 
            os.mkdir(Test_evol_PATH + "/" + cible + "/" + model + "/")
        for metric in metric_list[model]:
            if not os.path.exists(Test_evol_PATH + "/" + cible + "/" + model + "/" + metric): 
                os.mkdir(Test_evol_PATH + "/" + cible + "/" + model + "/" + metric)

if not os.path.exists(Test_evol_PATH + "/Patients/Comportement/"): 
    os.mkdir(Test_evol_PATH + "/Patients/Comportement/")
    
if not os.path.exists(Test_evol_PATH + "/Patients/Test_evolution_pval_patients.xlsx")  :
    wb = Workbook()
    wb.save(Test_evol_PATH + "/Patients/Test_evolution_pval_patients.xlsx")
    
if not os.path.exists(Test_evol_PATH + "/Controls/Test_evolution_pval_controls.xlsx")  :
    wb = Workbook()
    wb.save(Test_evol_PATH + "/Controls/Test_evolution_pval_controls.xlsx")
    """

#for cible in ["All"]:
for model in models:
    if not os.path.exists(Test_evol_PATH  + model + "/" ): 
        os.mkdir(Test_evol_PATH + model + "/")
    for metric in metric_list[model]:
        if not os.path.exists(Test_evol_PATH + model + "/" + metric): 
            os.mkdir(Test_evol_PATH +  model + "/" + metric)


if not os.path.exists(Test_evol_PATH + "/Test_evolution_diff.xlsx")  :
    wb = Workbook()
    wb.save(Test_evol_PATH + "/Test_evolution_diff.xlsx")

            
#%%
#cible = "Patients"
#models = ["dti"]
#region_list = ["CC", "UF"]
df_abs_chang_p = pd.DataFrame(region_list, columns=(["Region"]))
df_abs_chang_c = pd.DataFrame(region_list, columns=(["Region"]))
df_mean_c = pd.DataFrame(region_list, columns=(["Region"]))
df_mean_p = pd.DataFrame(region_list, columns=(["Region"]))
df_diff = pd.DataFrame(region_list, columns=(["Region"]))


for model in models:
    for metric in metric_list[model]:
        for region in region_list:
            
            data_p = (datas_p[["wMean_"+ metric + "_"+ model+"_T1_"+region, "wMean_"+ metric + "_"+ model+"_T2_"+region]]).copy().dropna(axis = 0)
            data_c = (datas_c[["wMean_"+ metric + "_"+ model+"_T1_"+region, "wMean_"+ metric + "_"+ model+"_T2_"+region]]).copy().dropna(axis = 0)
            
            #for controls
            data_T1_c = np.asarray(data_c["wMean_"+ metric + "_"+ model+"_T1_"+region])
            data_T2_c = np.asarray(data_c["wMean_"+ metric + "_"+ model+"_T2_"+region])
            
            chang_c = (data_T2_c - data_T1_c)
            abs_chang_c = np.mean(abs(chang_c))
            df_abs_chang_c.loc[df_abs_chang_c["Region"] == region,[model + "_" + metric]] = abs_chang_c
            mean_chang_c = np.mean(chang_c)
            df_mean_c.loc[df_mean_c["Region"] == region,[model + "_" + metric]] = mean_chang_c
            
            #for patients
            data_T1_p = np.asarray(data_p["wMean_"+ metric + "_"+ model+"_T1_"+region])
            data_T2_p = np.asarray(data_p["wMean_"+ metric + "_"+ model+"_T2_"+region])
            
            chang_p = (data_T2_p - data_T1_p)
            abs_chang_p = np.mean(abs(chang_p))
            df_abs_chang_p.loc[df_abs_chang_p["Region"] == region,[model + "_" + metric]] = abs_chang_p
            mean_chang_p = np.mean(chang_p)
            df_mean_p.loc[df_mean_p["Region"] == region,[model + "_" + metric]] = mean_chang_p
            
            #Comparison
            if abs_chang_p > abs_chang_c : 
                df_diff.loc[df_diff["Region"] == region,[model + "_" + metric]] = "significant"
                #print(metric, model, region, "diff significative")
            else :
                df_diff.loc[df_diff["Region"] == region,[model + "_" + metric]] = "NON sign."
                #print(metric, model, region, "diff NON significative")
                
            #ttest = stats.ttest_ind(chang_c, chang_p, equal_var = False)
            #df_pval.loc[df_pval["Region"] == region,[model + "_" + metric]] = ttest[1]
            a = """
            fig, ax = plt.subplots()      
            sns.boxplot([abs(chang_c), abs(chang_p)], palette = ["cadetblue","indianred"])
            ax.set_xticks([y for y in range(2)], labels=['Controls', 'Patients'])
            plt.title("Absolute percentage of change for\n" + metric + " " + model + " in "  + region)
            plt.savefig(Test_evol_PATH+ "/All/" + model + "/" + metric + "/Comparison_abs_%chang_" + model + "_" + metric + "_" + region + ".png")
            plt.show()"""

            

with pd.ExcelWriter(Test_evol_PATH+ "/Test_evolution_diff.xlsx",
                engine='openpyxl',mode='a', if_sheet_exists = 'replace') as writer:  
    df_abs_chang_p.to_excel(writer, sheet_name="Abs_diff_patients")
    df_abs_chang_c.to_excel(writer, sheet_name="Abs_diff_controls")
    df_mean_c.to_excel(writer, sheet_name="Mean_diff_patients")
    df_mean_p.to_excel(writer, sheet_name="Mean_diff_controls")
    df_diff.to_excel(writer, sheet_name="Comparison")
    
    